from django import http
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.contrib.auth.models import User
from .models import Type
from django.http import HttpResponse  
from app.functions.functions import handle_uploaded_file  
from .forms import StudentForm   
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm
from .models import InputForm
from django.contrib.auth import authenticate,login
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.template import RequestContext
 # Import mimetypes module
import mimetypes
# import os module
import os
# Import HttpResponse module
from django.http.response import HttpResponse
from django.http import HttpResponse  
from app.functions.functions import handle_uploaded_file  
from .forms import StudentForm   
import openpyxl
from .models import *


def AccueilDE(request):
    try:
        s=Fichiers.objects.get(type="salle_siq")
        sallesiq=getattr(s,"filename")   
      
    except Fichiers.DoesNotExist:
        sallesiq=None

    try:
        s=Fichiers.objects.get(type="salle_sit")
        sallesit=getattr(s,"filename")   
      
    except Fichiers.DoesNotExist:
        sallesit=None    
    
    try:
        s=Fichiers.objects.get(type="surv")
        surv=getattr(s,"filename")   
      
    except Fichiers.DoesNotExist:
        surv=None

    try:
        siq=Specialite.objects.get(titre="SIQ")
        canSiq=ListCandidats.objects.get(idSpecialite=siq)
        nomfSIQ = getattr(canSiq, "nomFichier")   
      
    except ListCandidats.DoesNotExist:
        canSiq=None
        nomfSIQ = None 

    try:
        sit=Specialite.objects.get(titre="SIT")
        canSit=ListCandidats.objects.get(idSpecialite=sit)
        nomfSIT = getattr(canSit, "nomFichier")

    except ListCandidats.DoesNotExist:
        canSit=None
        nomfSIT = None        
      

    return render(request,'de/AccueilDE.html', 
    {"sallesiq":sallesiq,"sallesit":sallesit,"surv":surv,
    "nomfSIQ":nomfSIQ,"nomfSIT":nomfSIT})



def affecterSalleSIQ(request) :   
    if "GET" == request.method:
        return redirect('AccueilDE')  
    else:   
        file = request.FILES["SIQ"]
        f=Fichiers(filename=file,type="salle_siq")
        f.save()
        wb = openpyxl.load_workbook(file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Feuil1"]
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        siq=Specialite.objects.get(titre="SIQ")    
        for i in range(len(excel_data)):
            Candidat.objects.filter(matricule=int(excel_data[i][0]),nom=excel_data[i][1],prenom=excel_data[i][2],dateNaiss=excel_data[i][3],specialite=siq).update(salle=excel_data[i][4],NumeroTable=excel_data[i][5])
    
        return redirect('AccueilDE')

def affecterSalleSIT(request) :   
    if "GET" == request.method:
        return redirect('AccueilDE')  
    else:
        file = request.FILES["SIT"]
        f=Fichiers(filename=file,type="salle_sit")
        f.save()
        wb = openpyxl.load_workbook(file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Feuil1"]
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        sit=Specialite.objects.get(titre="SIT")    
        for i in range(len(excel_data)):
            Candidat.objects.filter(matricule=int(excel_data[i][0]),nom=excel_data[i][1],prenom=excel_data[i][2],dateNaiss=excel_data[i][3],specialite=sit).update(salle=excel_data[i][4],NumeroTable=excel_data[i][5])

        return redirect('AccueilDE')

def affecterSurveillant(request) :   
    if "GET" == request.method:
        return redirect('AccueilDE')  
    else:
       
        file = request.FILES["Surveillant"]
        wb = openpyxl.load_workbook(file)
        f=Fichiers(filename=file,type="surv")
        f.save()
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Feuil1"]
        excel_data = list()
        nb_total=0
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            nb_total=nb_total+1
        
        request.session['excel'] = excel_data
        return redirect("AccueilDE")



def releverAbscence(request):
    if "GET" == request.method:
        
        data_siq=list()
        data_sit=list()
        siq=Specialite.objects.get(titre="SIQ")
        nbsiq=0
        nbsit=0
        try :
            cansiq=Candidat.objects.filter(specialite=siq,exclu=False)
            nbsiq=cansiq.count()
        except Candidat.DoesNotExist:
            cansiq=None
        sit=Specialite.objects.get(titre="SIT")
        try :
            cansit=Candidat.objects.filter(specialite=sit,exclu=False)
            nbsit=cansit.count()
        except Candidat.DoesNotExist:
            cansit=None
        
       
        return render(request, 'de/ReleverAbscence.html', {"data_siq":cansiq,"data_sit":cansit,"nbsiq":nbsiq,"nbsit":nbsit})
    else :  
        selected_values = request.POST.getlist('absent') 
        Candidat.objects.filter(matricule__in=selected_values).update(exclu=True)
        return redirect('releverAbscence')


def downloadListCan(request):
    if request.method == 'POST':  
        # Define Django project base directory
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # Define text file name
        if 'SIQ' in request.POST :
            siq=Specialite.objects.get(titre="SIQ")
            try:
                canSiq=ListCandidats.objects.get(idSpecialite=siq)
                filename = getattr(canSiq, "nomFichier") 
            except ListCandidats.DoesNotExist:
                filename = None       
           
        elif 'SIT' in request.POST  :
            sit=Specialite.objects.get(titre="SIT")
            try:
                canSit=ListCandidats.objects.get(idSpecialite=sit)
                filename = getattr(canSit, "nomFichier") 
            except ListCandidats.DoesNotExist:
                filename = None  
        if filename is None :
            return redirect('AccueilDE')
        else :    
            # Define the full file path
            filepath = BASE_DIR + '/app/static/upload/' + filename
            # Open the file for reading content
            path = open(filepath, 'r' ,encoding="cp437")
            # Set the mime type
            mime_type, _ = mimetypes.guess_type(filepath)
            # Set the return value of the HttpResponse
            response = HttpResponse(path, content_type=mime_type)
            # Set the HTTP header for sending to browser
            response['Content-Disposition'] = "attachment; filename=%s" % filename
            # Return the response value
            return response 
   



