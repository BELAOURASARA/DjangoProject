from django import http
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.contrib.auth.models import User
from .models import Type
from django.http import HttpResponse  
from app.functions.functions import handle_uploaded_file  
from .forms import StudentForm   
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm
from .models import InputForm
from django.contrib.auth import authenticate,login
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.template import RequestContext
 # Import mimetypes module
import mimetypes
# import os module
import os
# Import HttpResponse module
from django.http.response import HttpResponse
from django.http import HttpResponse  
from app.functions.functions import handle_uploaded_file  
from .forms import StudentForm   
import openpyxl
from .models import *

def AccueilDPGR(request):
    try:
        siq=Specialite.objects.get(titre="SIQ")
        canSiq=ListCandidats.objects.get(idSpecialite=siq)
        nomfSIQ = getattr(canSiq, "nomFichier")   
      
    except ListCandidats.DoesNotExist:
        canSiq=None
        nomfSIQ = None 

    try:
        sit=Specialite.objects.get(titre="SIT")
        canSit=ListCandidats.objects.get(idSpecialite=sit)
        nomfSIT = getattr(canSit, "nomFichier")

    except ListCandidats.DoesNotExist:
        canSit=None
        nomfSIT = None        

    try:
        correcteurs=Correcteur.objects.all()
        
        data=list()
        for i in range(len(correcteurs)):
            row_data = list()
            id=correcteurs[i].id
            user=User.objects.get(id=id)
            row_data.append(str(getattr(user, "first_name")))
            row_data.append(str(getattr(user, "last_name")))
            row_data.append(str(getattr(user, "email")))
            row_data.append(str(getattr(correcteurs[i], "epreuve")))
            data.append(row_data)
            
        
    except  Correcteur.DoesNotExist:
        correcteurs=None
    
    #return HttpResponse(data)
    return render(request,'dpgr/AccueilDPGR.html', 
    {"canSiq":canSiq,"canSit":canSit,"nomfSIQ":nomfSIQ,"nomfSIT":nomfSIT,"correcteurs":correcteurs,"data":data})


def ImporterCanSIQ(request):
    if request.method =="POST" :
        excel_file = request.FILES["SIQ"]
        
        #entrer le nom du fichier dans la base
        siq=Specialite.objects.get(titre="SIQ")
        listeCan=ListCandidats(nomFichier=excel_file,idSpecialite=siq)
        listeCan.save()
        #upload file
        handle_uploaded_file(excel_file)  
        #extract information from file
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Feuil1"]
        excel_data = list()
        
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        for i in range(len(excel_data)):
            can=Candidat(matricule=int(excel_data[i][0]),nom=excel_data[i][1],prenom=excel_data[i][2],dateNaiss=excel_data[i][3],salle=None,NumeroTable=None,exclu=False,specialite=siq)       
            can.save()
        
        return redirect('AccueilDPGR')   
    else :
        return redirect('AccueilDPGR')    
        
def ImporterCanSIT(request):
    if request.method =="POST" :
        excel_file = request.FILES["SIT"]
        #entrer le nom du fichier dans la base
        sit=Specialite.objects.get(titre="SIT")
        listeCan=ListCandidats(nomFichier=excel_file,idSpecialite=sit)
        listeCan.save()
        #upload file
        handle_uploaded_file(excel_file)  
        #extract information from file
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Feuil1"]
        excel_data = list()
        
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        for i in range(len(excel_data)):
            can=Candidat(matricule=int(excel_data[i][0]),nom=excel_data[i][1],prenom=excel_data[i][2],dateNaiss=excel_data[i][3],salle=None,NumeroTable=None,exclu=False,specialite=sit)       
            can.save()
        
        return redirect('AccueilDPGR')   
    else :
        return redirect('AccueilDPGR') 

def ImporterCorriges(request):           
    if request.method =="POST" :
        epreuve=request.POST.get("epreuve")
        excel_file = request.FILES["corrige"]
        #entrer le nom du fichier dans la base
        corr=CorrigesType(nomFichier=excel_file,Epreuve=epreuve)
        corr.save()
        #upload file
        handle_uploaded_file(excel_file)   
        return redirect('AccueilDPGR')   
    else :
        return redirect('AccueilDPGR')  

def ImporterEns(request):
    if request.method =="POST" :
        excel_file = request.FILES["filename"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Feuil1"]
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        for i in range(len(excel_data)):
            concat=excel_data[i][0]+excel_data[i][1]
            corr=User(first_name=excel_data[i][0],last_name=excel_data[i][1],email=excel_data[i][2],username=concat)       
            corr.set_password(concat)
            corr.save()
            correcteur=Correcteur(id=corr.id,epreuve=excel_data[i][3])
            correcteur.save()
            t=Type(id=corr.id,Type=2)
            t.save()
        return redirect('AccueilDPGR')  
    else :
        return redirect('AccueilDPGR')   

def dashboard_DPGR(request):
    return render(request,'dpgr/dashboard_dpgr.html',{})

def Mesdocuments_DPGR(request):
    try:
        corr=CorrigesType.objects.all() 
      
    except Fichiers.DoesNotExist:
        corr=None

    try:
        s=Fichiers.objects.get(type="salle_siq")
        sallesiq=getattr(s,"filename")   
      
    except Fichiers.DoesNotExist:
        sallesiq=None  
    
    try:
        s=Fichiers.objects.get(type="salle_sit")
        sallesit=getattr(s,"filename")   
      
    except Fichiers.DoesNotExist:
        sallesit=None    
    
    try:
        s=Fichiers.objects.get(type="surv")
        surv=getattr(s,"filename")   
      
    except Fichiers.DoesNotExist:
        surv=None

    try:
        siq=Specialite.objects.get(titre="SIQ")
        canSiq=ListCandidats.objects.get(idSpecialite=siq)
        nomfSIQ = getattr(canSiq, "nomFichier")   
      
    except ListCandidats.DoesNotExist:
        canSiq=None
        nomfSIQ = None 

    try:
        sit=Specialite.objects.get(titre="SIT")
        canSit=ListCandidats.objects.get(idSpecialite=sit)
        nomfSIT = getattr(canSit, "nomFichier")

    except ListCandidats.DoesNotExist:
        canSit=None
        nomfSIT = None        
      

    return render(request,'dpgr/Mesdocuments_DPGR.html', 
    {"corr":corr,"sallesiq":sallesiq,"sallesit":sallesit,"surv":surv,
    "nomfSIQ":nomfSIQ,"nomfSIT":nomfSIT})
    

 
   
